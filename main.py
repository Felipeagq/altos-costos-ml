from datetime import datetime
import glob
import os
import procesador
import zipfile
from openpyxl import Workbook
from openpyxl import load_workbook
import boto3
import json
import logging
from botocore.exceptions import ClientError
import time
from pdf_to_txt1 import pdf_to_txt
from dotenv import load_dotenv
# Importamos la libreria
import fitz
import cv2
import os
import sys
import pandas as pd
from PIL import Image
import pytesseract
import re


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract'

load_dotenv()

# Creamos el cliente s3
ACCESS_KEY_ID = os.getenv("ACCESS_KEY_ID")
SECRET_ACCESS_KEY = os.getenv("SECRET_ACCESS_KEY")


def normalize(s):
    de = ['MOTIVO DE CONSULTA',
        'ENFERMEDAD ACTUAL',
        'EXAMEN',
        'ANÁLISIS',
        'PLAN Y MANEJO',
        'DIAGNÓSTICO',
        'ORDENES',
        'ÓRDENES',
        'INTERCONSULTAS',
        'NOTAS ENFERMERIA',
        'FORMULA MÉDICA',
        'FORMATOS',
        'RECOMENDACIONES',
        'NOTA DE INGRESO',
        'OBSERVACIONES']
    a = ['$$MOTIVO DE CONSULTA',
        '$$ENFERMEDAD ACTUAL',
        '$$EXAMEN',
        '$$ANÁLISIS',
        '$$PLAN Y MANEJO',
        '$$DIAGNÓSTICO',
        '$$ORDENES',
        '$$ÓRDENES',
        '$$INTERCONSULTAS',
        '$$NOTAS ENFERMERIA',
        '$$FORMULA MÉDICA',
        '$$FORMATOS',
        '$$RECOMENDACIONES',
        '$$NOTA DE INGRESO',
        '$$OBSERVACIONES']
    replacements = (zip(de,a) )
    for a, b in replacements:
        s = s.replace(a, b)
    return s

# Función para crear el LINK de descarga del xlsx
def create_presigned_url(cliente, bucket_name, object_name, expiration=604800):
    """Generate a presigned URL to share an S3 object

    :param bucket_name: string
    :param object_name: string
    :param expiration: Time in seconds for the presigned URL to remain valid
    :return: Presigned URL as string. If error, returns None.
    """
    # Creamos el link de descarga
    try:
        response = cliente.generate_presigned_url('get_object',
                                                    Params={'Bucket': bucket_name,
                                                            'Key': object_name},
                                                    ExpiresIn=expiration)
    except ClientError as e:
        logging.error(e)
        return None
    # The response contains the presigned URL
    return response

fcorte1 = ''
eps1 = ''

# Función principal
def main(fecha,eps,cliente_mqtt,hash):
    # Enumeramos todos los archivos .pdf
    hcs = glob.glob('*.pdf')
    n = (len(hcs))*2 +1
    print(f'El numero de total de procentajes: {n}')
    actual = 1 
    data = json.dumps({
        "hash":hash,
        "step" : "processing",
        "percentage":'0%',
        "page": " ",
        "link": " "
    })
    cliente_mqtt.publish(hash,data)    
    for hc in hcs:
        print(hc)
        # convertimos los .pdf en .txt
        #pdf_to_txt(hc[:-4],cliente_mqtt)
        procentaje = f'{round(((actual/n) *100),3)}%'
        data = json.dumps({
            "hash":hash,
            "step" : "processing",
            "percentage":procentaje,
            "page": " ",
            "link": " "
        })
        cliente_mqtt.publish(hash,data)
        #print(hc[:-4])
        print('comenzó pdf_to_txt')
        # Ruta del documento de pdf
        pdf_documento = '{}.pdf'.format(hc[:-4])
        print('pdf_documento: ', pdf_documento)
        # Creamos el objeto del documento de pdf abierto
        pdf_documento = fitz.open(pdf_documento)
        print('pdf_documento: ', pdf_documento)
        # Guardamos la pagina 0 como imagen 
        texto = []
        print('texto: ', texto)
        documento = fitz.open(pdf_documento)
        print('documento: ', documento)
        page = documento[0]
        print('page_1: ', page)
        pix = page.get_pixmap()
        print('pix: ', pix)
        print('page.number: ', page.number)
        pix.pil_save("page-{}.jpg".format(page.number))
        print("pix.pil_save: ", pix.pil_save("page-{}.jpg".format(page.number)))
        imagen = cv2.imread('page-{}.jpg'.format(page.number))
        # print('imagen: ', imagen)
        print('imagen.shape[1]: ', imagen.shape[1])
        print('imagen.shape[0]: ', imagen.shape[0])
        (x,y,w,h,x2,y2) = 0,0, imagen.shape[1], 245, imagen.shape[1], imagen.shape[0]
        print((x,y,w,h,x2,y2))
        superior = imagen[y:y+h,x:x+w]
        # print('superior: ', superior)
        cv2.imwrite('superior.jpg',superior)

        imagen = cv2.imread('superior.jpg')

        # print('imagen superior: ', imagen)

        imagen = cv2.resize(imagen,(1268,460)) 

        # print('imagen con dimensiones cambiadas: ', imagen)
        print('page.number: ', page.number)

        # print('current_directory: ', os.getcwd())

        text = pytesseract.image_to_string(Image.open('page-{}.jpg'.format(page.number)),lang='spa')

        # print('text: ', text)

        paciente = text[text.index('HISTORIA CLÍNICA No.'):text.index('HISTORIA CLÍNICA No.') + text[text.index('HISTORIA CLÍNICA No.'):].index('\n')]
        
        # print('0paciente:', paciente)

        texto = text
        #texto = texto.lower()
        texto = texto.replace("Afiliado","\nAfiliado")
        texto = texto.replace("Edad actual","\nEdad actual")
        texto = texto.replace("Sexo","\nSexo")
        texto = texto.replace("Grupo Sanguíneo","\nGrupo Sanguíneo")
        texto = texto.replace("Estado Civil","\nestado civil")
        texto = texto.replace("Dirección","\nDirección")
        texto = texto.replace("Departamento","\nDepartamento")
        texto = texto.replace("Ocupacion","\nOcupacion")
        texto = texto.replace("Grupo Etnico","\nGrupo Etnico")
        texto = texto.replace("Atención Especial","\nAtención Especial")
        texto = texto.replace("\n\nDiscapacidad","\nDiscapacidad")
        texto = texto.replace("Grupo Poblacional","\nGrupo Poblacional")
        
        # print('el texto que se muestra es: ', texto)
        print('1paciente: ', paciente)

        file = open(f"{paciente}.txt","w")
        file.write(texto)
        file.close()

        paginas = len(documento)

        print('paginas: ', paginas)
        print('documento: ', documento)

        for pagina in range(0,paginas):
            page = documento[pagina]
            # print('page_2: ', page)
            text = page.get_text("text")
            # print('type(texto): ', type(texto))
            texto.join(text)
            pagina_actual = page.number + 1
            print('pagina_actual: ', pagina_actual)
            data_step = json.dumps({
                "hash":hash,
                "step" : "processing",
                "percentage":procentaje,
                "page" : f"{pagina_actual}/{paginas}",
                "link": " "
            })
            # print('data_step: ', data_step)        
            # print(f"Vamos por la pagina {pagina_actual} / {paginas}")
            cliente_mqtt.publish(hash,data_step)
            if os.path.exists("page-{}.png".format(page.number)):
                os.remove("page-{}.png".format(page.number))

        if os.path.exists("inferior.jpg"):
            os.remove("inferior.jpg")
        if os.path.exists("superior.jpg"):
            os.remove("superior.jpg")

        file = open(f"{paciente}.txt","a")
        for text in texto:
            text = normalize(text)
            file.write(text.replace('FOLIO','==> FOLIO'))
        file.close()
        actual = actual + 1
    # Enumeramos todos los archivos 'HISTORIA*.txt'
    pacientes = glob.glob('HISTORIA*.txt')
    row = 7
    # Procesamos cada 'HISTORIA*.txt' y lo colocamos en el xlsx
    fcorte1 = fecha
    eps1 = eps
    print('pacientes:', pacientes)
    try:
        for paciente in pacientes:
            try:
                print('2paciente:', paciente)
                procesador.main(paciente,row,fecha,eps)
            except Exception as e:
                print('La excepcion que se da es:', e)
                print(f"el proceso de {paciente} explota")
            # print('row:', row)
            print('3paciente:', paciente)
            time.sleep(0.5)
            procentaje = f'{((actual/n) *100)}%'
            data = json.dumps({
            "hash":hash,
            "step" : "processing",
            "percentage":procentaje,
            "page": "done",
            "link": " "
            })
            actual = actual + 1
            print('actual: ', actual)
            cliente_mqtt.publish(hash,data)
            # aqui va el MQTT
            row = row + 1
            print('-- -- -- -- -- -- -- -- -- -- -- -- -- -- ')
    except Exception as e:
        print('el error de la excepcion es:', e)

    # Eliminamos los .pdf del local
    for hc in hcs:
        try:
            os.remove(hc)
            cliente_mqtt.publish(hash,hc)
            print(f"se eliminó {hc}")
        except:
            continue
    
    # creamos el cliente s3
    s3_client = boto3.client("s3",
                        region_name="us-east-1",
                        aws_access_key_id=ACCESS_KEY_ID,
                        aws_secret_access_key=SECRET_ACCESS_KEY
                        )
    
    # Subimos el archivo .pdf al s3
    with open("prueba2.xlsx","rb") as archivo_file:
        data = archivo_file.read()
    # definimos los parametros de subida
    rigth_now = datetime.now()
    timestamp = datetime.timestamp(rigth_now)
    response = s3_client.put_object(
        Body= data,
        Bucket = "macna-data",
        Key= f"pdf-upload/,/MACNA-{eps}-{fecha}-{int(timestamp)}.xlsx"
    )
    print('response:', response)
    archivo_file.close()
    # Generamos el link de descarga del xlsx
    print("generamos link")
    link =  create_presigned_url(s3_client,"macna-data",f"pdf-upload/,/MACNA-{eps}-{fecha}-{int(timestamp)}.xlsx")

    print('link:', link)
    
    time.sleep(1)
    
    ##Borramos las modificaciones realizadas en el .xlsx
    # row = 7
    # wb = load_workbook(filename="prueba2.xlsx")
    # ws = wb['CAC']
    # for i in range(n):
    #     # for i in range(1,167):
    #     #     try:
    #     #         # Borramos la celda.
    #     #         ws.cell(row=row,column=i,value=" ")
    #     #     except:
    #     #         continue
    #     row = row + 1
    # # Guardamos los cambios realizados
    # wb.save("prueba2.xlsx")
    
    ##Eliminamos los 'HISTORIA*.txt' generados
    for paciente in pacientes:
        try:
            os.remove(paciente)
        except:
            continue
    
    # Retornamos el link de descarga del .xlsx
    return link 
    
if __name__=='__main__':
    print("esta corriendo como __main__")
