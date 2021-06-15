import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
excel_1 = 'prueba2.xlsx'
excel_2 = "prueba.xlsx"


#wb = Workbook() # creamos objeto de Excel
#wb.save(excel_1) 
wb = load_workbook(filename=excel_2)
#wb.create_sheet('CAC',0)
ws = wb['CAC']
for i in range(1,167):
    try:
        ws.cell(row=1,column=i,value='hola')
    except:
        continue

wb.save(excel_2)