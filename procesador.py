__1 = None
__2 = None
__3 = None
__4 = None
__5 = 'NONE'
__6 = None
__7 = None
__8 = None
__9 = None
__10 = None
__11 = None
__12 = '6'
__13 = None
__14 = None
__15 = None
__16 = '1800-01-01' # PREDETERMINADA
__17 = None
__18 = None
__19 = ' '
__20 = ' '
__21 = ' '
__22 = ' ' # PREDETERMINADA
__23 = " "
__24 = " "
__25 = ' ' # 80010054401
__26 = " "
__27 = " "
__28 = " "
__29 = ' '
__30 = ' '
# MAMA
__31 = '98'
__32 = '1845-01-01'
__33 = '98'
__34 = '98'
__35 = '1845-01-01'
__36 = '98'
__37 = '98'
__38 = '98'
__39 = '1845-01-01'
__40 = '1'
__41 = '3'
__42 = '2'
__43 = '1845-01-01'
__44 = '99'
__45 = '98'
__46 = '98'
__47 = '97'
__48 = '97'
__49 = '97'
__50 = '97'
__51 = '97'
__52 = '97'
__53 = '97'
__54 = '97'
__55 = '98'
__56 = ' '
__57 = '1845-01-01'
__58 = '98'
__59 = '98'
__60 = '98'
__61 = '98'
__62 = '98'
__63 = '98'
__64 = '98'
__65 = '98'
__66 = '98'
__67 = '98'
__68 = '98'
__69 = '98'
__70 = '98'
__71 = '98'
__72 = '98'
__73 = '98'
__74 = '98'
__75 = '1845-01-01'
__76 = '98'
__77 = '98'
#__78 = '98'
__78 = " "
__79 = '1845-01-01'
__80 = '98'
__81 = '98'
__82 = '98'
__83 = '98'
__84 = '98'
__85 = '98'
__86 = '98'
__87 = '98'
__88 = '98'
__89 = '98'
__90 = '98'
__91 = '98'
__92 = '98'
__93 = '98'
__94 = '98'
__95 = '98'
__96 = '98'
__97 = '1845-01-01'
__98 = '98'
__99 = '98'
__100 = "98"
__101 = '98'
__102 = '1845-01-01'
__103 = '98'
__104 = '98'
__105 = '98'
#__105 = " "
__106 = '1845-01-01'
__107 = '98'
__108 = '98'
__109 = '98'
__110 = ' '
__111 = '98'
__112 = '98'
__113 = '98'
__114 = '1845-01-01'
__115 = ' '
__116 = '98'
__117 = '98'
__118 = '98'
__119 = '98'
__120 = '1845-01-01'
__121 = '98'
__122 = '98'
__123 = '1845-01-01'
__124 = ' '
__125 = '98'
__126 = '98'
__127 = '98'
__128 = '98'
__129 = '1845-01-01'
__130 = '98'
__131 = '98'
__132 = '98'    # PREDETERMINADA
__133 = '98'    # PREDETERMINADA
__134 = '98'    # PREDETERMINADA
__135 = '1845-01-01'    # PREDETERMINADA
__136 = '98'    # PREDETERMINADA
__137 = '98'    # PREDETERMINADA
__138 = '1845-01-01'    # PREDETERMINADA
__139 = '98'    # PREDETERMINADA
__140 = '2'
__141 = '2'
__142 = '2'
__143 = '2'
__144 = '2'
__145 = '2'
__146 = ' '
__147 = '1845-01-01'
__148 = '98'
__149 = '98'    # PREDETERMINADA
__150 = '1845-01-01'    # PREDETERMINADA
__151 = '98'    # PREDETERMINADA
__152 = '98'
__153 = '1845-01-01'
__154 = '98'
__155 = '4'
__156 = '98' # PREDETERMINADA
__157 = '9'
__158 = '5'
__159 = '1'
__160 = '0'
__161 = '1'
__162 = '1845-01-01'    # PREDETERMINADA
__163 = '1845-01-01'
__164 = '98'
__165= '98' # PREDETERMINADA
__166 = '2021-01-01'

import time 
from pdf_to_txt1 import pdf_to_txt
import glob
import os
import requests
import json
import pandas as pd
from openpyxl import Workbook
import re
from openpyxl import load_workbook


#import nltk
#import cac_data_4
def main(Paciente,row,Fcorte,Eps):
    #Convertimos el pdf en texto
    #pdf_to_txt_3.pdf_to_csv('CC 40925684 - HC')

    # Cargamos el diccionario
    file = open("dic_1.json",)
    dic = json.load(file)

    ####################################
    ### DECLARAMOS ALGUNAS FUNCIONES ###
    ####################################

    # Funcion para quitar los acentos 
    def mayusculas_sec_clave(s):
        ''' Funcion para volver mayusculas las secciones clave'''
        replacements =(
            ('motivo de consulta','MOTIVO DE CONSULTA'),
            ('enfermedad actual','ENFERMEDAD ACTUAL'),
            ('examen fisico','EXAMEN FÍSICO'),
            ('analisis','ANÁLISIS'),
            ('evolucion medico','EVOLUCION MEDICO'),
            ('antecedentes','ANTECEDENTES'),
            ('plan y manejo','PLAN Y MANEJO'),
            ('diagnostico','DIAGNÓSTICO'),
            ('diagnóstico','DIAGNÓSTICO'),
            ('ordenes','ORDENES'),
            ('órdenes','ÓRDENES'),
            ('interconsultas','INTERCONSULTAS'),
            ('notas enfermeria','NOTAS ENFERMERIA'),
            ('formula médica','FORMULA MÉDICA'),
            ('formatos','FORMATOS'),
            ('evolucion medico','EVOLUCION MEDICO'),
            ('recomendaciones','RECOMENDACIONES'),
            ('descripcion cirugia',"DESCRIPCIÓN CIRUGÍA"),
            ('nota de ingreso','NOTA DE INGRESO'),
            ('observaciones','OBSERVACIONES')
            )
        for a, b in replacements:
            s = s.replace(a, b)
        return s

    def lower(text):
        return text.lower()

    def normalize(s):
        replacements = (
            ("á", "a"),
            ("é", "e"),
            ("í", "i"),
            ("ó", "o"),
            ("ó", "o"),
            ("ú", "u"),
            ('\xa0','')
        )
        for a, b in replacements:
            s = s.replace(a, b)
        return s

    def aux(text,search,helper=False):
        start = text.find(search)
        aux = text[start:]
        end = aux.find('\n')
        fragmento = aux[:end]
        if helper==True:
            fragmento = fragmento[fragmento.find(':')+2:]
            return fragmento.replace('\n','')
        return fragmento.replace('\n','')


    def search_text_in_text(texto,dic,tipo):
        success = 0
        for item in dic[tipo].keys():
            if item in texto:
                success = 1
            if texto.count(item) != 0:
                print(f'{item} --> {dic[tipo][item][0]} ')
        if success == 0:
            print(dic[tipo]['na'])

    def aux_2(text,search):
        try:
            start = text.find(search)
            aux = text[start+3:]
            end = aux.find('$ $')
            fragmento = aux[:end]
            return fragmento
        except:
            fragmento = None
            return fragmento


    def aux_reg(text,search):
        try:
            start = text.find(search)
            aux = text[start:]
            end = aux.find('sede')
            if 'sede' not in aux:
                end = aux.find('75.0 *hosvital*')
            fragmento = aux[:end]
            return fragmento
        except:
            fragmento = None
            return fragmento

    def findKeyWord(text,keyWord,n):
        wordlist = text.split()
        here = []
        for i in range(len(wordlist)):
            if wordlist[i] == keyWord:
                if i > n:
                    nGram = wordlist[i-n:i+n+1]
                elif i+n > len(wordlist):
                    nGram = wordlist[i-n:]
                elif ((2*n)+1)>len(wordlist):
                    nGram = wordlist[:]
                else:
                    nGram = wordlist[:i+n]
                here.append(nGram)
        return here



    ###########################################
    ### EMPEZAMOS EL PROCESAMIENTO GENTERAL ###
    ###########################################
    secciones = []
    with open(Paciente,'r', encoding = 'utf-8', errors='ignore') as file:
        paciente = file.read()
        paciente = paciente.lower().replace('\n\n',' \n\n ') # quitamos los "\n"
        paciente = mayusculas_sec_clave(paciente) # agregamos las mayusculas de las secciones
        paciente = normalize(paciente) # quitamos los acentos de la data
        #paciente = remove_punctuation(paciente)
        folios = paciente.split('==>') # separamos por secciones 

        # print('paciente_procesador1:', paciente )
            


    #with open("HISTORIA CLÍNICA No.CC 84046881 -- WARNER CHISTIAN MONTENEGRO BARROS.txt",'r') as file:
    #    paciente = file.read()
    #paciente = paciente.lower()
    #paciente = mayusculas_sec_clave(paciente)
    #paciente = normalize(paciente)



    # Sacamos la informacion del encabezado
    def info_Encabezado(texto,Eps):
        paciente1 = ''
        '''
        Esta función extrae la información de los items
        1 --> 16
        los cuales aparecen en el encabezado del historial clinico
        -----------
        Parametros:
        -----------
        - texto (string)    :el historial clinico convertido en texto
                            y normalizado.

        '''
        __1 = None
        def lower(text):
            return text.lower()
        def normalize(s):
            replacements = (
                ("á", "a"),
                ("é", "e"),
                ("í", "i"),
                ("ó", "o"),
                ("ó", "o"),
                ("ú", "u"),
            )
            for a, b in replacements:
                s = s.replace(a, b)
            return s
        
        head = aux(texto,'historia clinica')
        print('head:', head)
        # identificacion
        head2 = head[head.find('.')+1:head.find('.')+3]
        name = head[head.find('- ')+2:].split()

        global __5
        try:
            __5 = head2.upper()
        except Exception as e:
            print('La excepcion de __5 es:', e)
        if __5 is not None:
            __5 = 'CC'

        print('__5:', __5)

        global __6
        a = re.findall('[0-9]+', head)
        try:
            __6 = a[0]
        except Exception as e:
            print('La excepcion de __6 es:', e)
        if __6 is None:
            __6= '0'

        print('__6:', __6)

        # nombres
        global __4

        try:
            header = {"X-Authorization":"OcUacy2Q3REsQX4KPA2x7LnMYrNo0HthgAIFt6YKYvuQNOSimUgzPGMcFyN376jJ"}
            res = requests.get(f"http://190.131.222.108:8088/api/v1/macna/patient/0/type/{__5}/information",headers=header)
            persona = json.loads(res.text)
            print('persona:', persona)
            paciente1 = persona["data"][0]
            __1 = paciente1.get("fName","NONE")          
            __2 = paciente1.get("sName","NONE")            
            __3 = paciente1.get("fLastname","NONE")
            __4 = paciente1.get("sLastname","NONE")
        # nombres
        except Exception as e:
            print('la excepcion para poner los headers es:', e)
        print('name:', name)
        if (len(name)) > 0:
            if name[0] is not None:
                __1 = name[0].upper()
            else:   
                __1 = "NONE"

            if name[1] is not None:
                __2 = name[0].upper()
            else:   
                __2 = "NONE"
            
            if name[2] is not None:
                __3 = name[2].upper()
            else:   
                __3 = "NONE"
            
            try:
                __4 = name[3].upper()
            except Exception as e:
                print('La excepcion de __4 es:', e)

            if __4 is None:
                __4 = "NONE"
        else: 
            __1 = "NONE"
            __2 = "NONE"
            __3 = "NONE"
            __4 = "NONE"
        print('__1:', __1)
        print('__2:', __2)
        print('__3:', __3)
        print('__4:', __4)

        # fecha nacimiento
        global __7
        try:
            fecha = paciente1.get("birthDate")
            __7 = fecha
        except Exception as e:
            print('La excepcion de __7 es:', e)

        print('__7:', __7)

        # sexo
        global __8
        try:
            __8 = paciente1.get("gender")
        except Exception as e:
            print('La excepcion de __8 es:', e)

        print('__8:', __8)

        edad = 0
        try:
            edad = paciente1.get("age")
            edad = int(edad)
        except Exception as e:
            print('La excepcion de edad es:', e)

        print('edad:', edad)

        if edad < 60:
            __9 = '9622'
        elif __8 == 'F' and edad > 60:
            __9 = '9111'
        else:
            __9 = '9629'
        
        if __5 == 'TI':
            __9 = '9998' 
        eps = aux(texto,'empresa:',True) #COMFAGUAJIRA PGP ONCOLOGIA y SUBSIDIADO son diferentes?
        __11 = Eps

        if 'subsidiado' in eps:
            __10 = 'S'
        else:
            __10 = 'C'
        # mirar el df[df['tipo']==11]['texto'].unique()
        #grupo2
        etnia = aux(texto,'etnia:',True)
        # PERTENENCIA ETNICA
        if etnia == 'indigena'.upper():
            __12 = '1'
        elif  etnia == 'ROM':
            __12 = '2'
        elif  etnia == 'Raizal'.upper():
            __12 = '3'
        elif etnia == 'palenquero'.upper():
            __12 = '4'
        elif  etnia == 'negro'.upper():
            __12 = '5'
        else:
            __12 = '6'
        if edad > 60:
            __13 = '31'
        else:
            __13 = '5'
        # residencia

        municipio = ''
        try:
            municipio = paciente1.get("city","BARRANQUILLA")
        except Exception as e:
            print('La excepcion de municipio es:', e)

        print('municipio:', municipio)
        
        data = pd.read_csv('municipios.csv',sep=',')
        try:
            codigo = data[data['municipio']==municipio]['codigo'].values[0]
        except:        
            codigo = "08001"
        print('codigo:', codigo)
        __14 = codigo

        global __15
        try:
            __15 = paciente1.get("mobilePhone","0")
        except Exception as e:
            print('La excepcion de __15 es:', e)
        if __15 is None:
            __15 = "NONE"
        __16 = "1800-01-01" # en todos los casos a sido na

        print('__15:', __15)

        return __1,__2,__3,__4,__5,__6,__7,__8,__9,__10,__11,__12,__13,__14,__15,__16
    
    global __1,__2,__3,__4,__5,__6,__7,__8,__9,__10,__11,__12,__13,__14,__15,__16
    __1,__2,__3,__4,__5,__6,__7,__8,__9,__10,__11,__12,__13,__14,__15,__16 = info_Encabezado(paciente,Eps)
    v_head = [__1,__2,__3,__4,__5,__6,__7,__8,__9,__10,__11,__12,__13,__14,__15,__16]

    print('v_head:', v_head)

    for i in range(1,17):
        try:
            wb = load_workbook(filename="prueba2.xlsx")
            ws = wb['CAC']
            ws.cell(row=row,column=i,value=v_head[i-1])
            wb.save("prueba2.xlsx")
        except:
            wb = load_workbook(filename="prueba2.xlsx")
            ws = wb['CAC']
            ws.cell(row=row,column=i,value="N/A")
            wb.save("prueba2.xlsx")

    # print('---------')    
    # print('wb: ', wb)
    # print('ws: ', ws)
    # print('---------')

    # VERSION DEL APLICATIVO #
    ##########################
    wb = load_workbook(filename="prueba2.xlsx")
    ws = wb['CAC']
    ws.cell(row=1,column=1,value="v1.1.2")
    wb.save("prueba2.xlsx")
    
    # print('wb:', wb)
    # print('ws:', ws)

    ###################
    ### VARIABLE 17 ###
    ###################


    def _17(folios,dic,__6,__5):
        import statistics
        from statistics import mode
               
        try:
            header = {"X-Authorization":"OcUacy2Q3REsQX4KPA2x7LnMYrNo0HthgAIFt6YKYvuQNOSimUgzPGMcFyN376jJ"}
            print('header:', header)
            res = requests.get(f"http://190.131.222.108:8088/api/v1/macna/patient/{__6}/type/{__5}/information",headers=header)
            persona = json.loads(res.text)
            diag = ['C']
            for i in range(len(persona["data"])):
                diag.append(persona["data"][i]["diagnostics_cod"])
            diag =  [x for x in diag if x is not None]
            print('diag: ', diag)
            diagx = statistics.mode(diag)
            print('diagx:', diagx)
            print("llega aca")
            if "C" in diagx:
                print("entra al if")
                return diagx
            print("entrando al for")
            for c in diag:
                print(c)
                if "C" in c:
                    return c
            print("termino el try")             
        except Exception as e:
            print('lo que esta ocasionando este problema es:', e)
            print("entrando en except")
            here = []
            diagnosticos = []
            for folio in range(len(folios)):
                if '$$DIAGNÓSTICO' in folios[folio]:
                    here.append(folio)
            # try:
                # preguntar si es el primer diagnostico
                # el ultimo diagnostico
                # o el más frecuente

            print('here:', here)
            
            for i in here:
                to_back = folios[i].find("$$DIAGNÓSTICO")
                diagnostic = folios[i][to_back-5:to_back]
                diagnostic = diagnostic.replace("\n","")

                print('diagnostic:', diagnostic)
                if not diagnostic.isalnum():
                    continue
                if "c" in diagnostic:
                    diagnosticos.append(diagnostic)
            print("Diagnostico encontrado")
            print('diagnosticos:', diagnosticos)
            diag = mode(diagnosticos)
            print('diag;', diag)

            global __17
            __17 = diag
            print('__17:', __17)
            return __17.replace("\n","").replace("c","C").upper()
            
            # except:
            #     return 'No hubo diagnostico'
    global __17
    if __17 is not None:
        __17 = _17(folios,dic,__6,__5)
        __17 = __17.upper()
    else:
        print("fallo __17")
        __17 = 'NONE'
    
    print('__17:', __17)







    import antecedentes
    print("-- -- -- ANTECEDENTES -- -- -- ")
    





    ###################
    ## QUIMIOTERAPIA ##
    ###################
    import quimioterapia
    print("-- -- -- QUIMIOTERAPIA -- -- -- ")
    global __18 ,__19 ,__20 ,__21 ,__22 ,__23 ,__24 ,__25 ,__26 ,__27 ,__28 ,__29 ,__30 ,__31 ,__32 ,__33 ,__34 ,__35 ,__36 ,__37 ,__38 ,__39 ,__40 ,__41 ,__42 ,__43 ,__44 ,__45 ,__46 ,__47 ,__48 ,__49 ,__50 ,__51 ,__52 ,__53 ,__54 ,__55 ,__56 ,__57 ,__58 ,__59 ,__60 ,__61 ,__62 ,__63 ,__64 ,__65 ,__66 ,__67 ,__68 ,__69 ,__70 ,__71 ,__72 ,__73 ,__74 ,__75 ,__76 ,__77 ,__78 ,__79 ,__80 ,__81 ,__82 ,__83 ,__84 ,__85 ,__86 ,__87 ,__88 ,__89 ,__90 ,__91 ,__92 ,__93 ,__94 ,__95 ,__96 ,__97 ,__98 ,__99 ,__100 ,__101 ,__102 ,__103 ,__104 ,__105 ,__106 ,__107 ,__108 ,__109 ,__110 ,__111 ,__112 ,__113 ,__114 ,__115 ,__116 ,__117 ,__118 ,__119 ,__120 ,__121 ,__122 ,__123 ,__124 ,__125 ,__126 ,__127 ,__128 ,__129 ,__130 ,__131 ,__132 ,__133 ,__134 ,__135 ,__136 ,__137 ,__138 ,__139 ,__140 ,__141 ,__142 ,__143 ,__144 ,__145 ,__146 ,__147 ,__148 ,__149 ,__150 ,__151 ,__152 ,__153 ,__154 ,__155 ,__156 ,__157 ,__158 ,__159 ,__160 ,__161 ,__162 ,__163 ,__164 ,__165 ,__166
    try:
        __45,__46,__47,__48,__49,__50,__51,__52,__53,__54,__55,__56,__57,__58,__59,__60,__61,__62,__63,__64,__65,__66,__67,__68,__69,__70,__71,__72,__73,__74,__75,__76,__77,__78,__79, __80, __81, __82, __83, __84, __85, __86, __87, __88, __89, __90, __91, __92, __93, __94, __95, __96, __97, __98, __99 = quimioterapia.main(__6,__5,__17)
        print(f""" 
        PRIMER CICLO DE QUIMIOTERAPIA:
        __45: {__45}
        __46: {__46}
        __47: {__47}
        __48: {__48}
        __49: {__49}
        __50: {__50}
        __51: {__51}
        __52: {__52}
        __53: {__53}
        __54: {__54}
        __55: {__55}
        __56: {__56}
        __57: {__57}
        __58: {__58}
        __59: {__59}
        __60: {__60}
        __61: {__61}
        __62: {__62}
        __63: {__63}
        __64: {__64}
        __65: {__65}
        __66: {__66}
        __67: {__67}
        __68: {__68}
        __69: {__69}
        __70: {__70}
        __71: {__71}
        __72: {__72}
        __73: {__73}
        __74: {__74}
        __75: {__75}
        __76: {__76}
        __77: {__77}
        SEGUNDO CICLO DE QUIMIOTERAPIA:
        __78: {__78}
        __79: {__79}
        __80: {__80}
        __81: {__81}
        __82: {__82}
        __83: {__83}
        __84: {__84}
        __85: {__85}
        __86: {__86}
        __87: {__87}
        __88: {__88}
        __89: {__89}
        __90: {__90}
        __91: {__91}
        __92: {__92}
        __93: {__93}
        __94: {__94}
        __95: {__95}
        __96: {__96}
        __97: {__97}
        __98: {__98}
        __99: {__99}
                """)
    except Exception as e:
        print(e)
        print("falló _45__77_")

    print("-- -- -- -- -- -- -- -- -- -- -- --")

    
    ###############
    ### CIRUGIA ###
    ###############
    import cirugias_peticion

    global __100,__101,__102,__103,__104,__105,__106,__107,__108,__109,__110,__111

    try:
        __100,__101,__102,__103,__104,__105,__106,__107,__108,__109,__110,__111 = cirugias_peticion.main(__6,__5)
    except Exception as e:
        print(e)
        print("falló _100__111")

    print("100 : ",__100)

    ###################################
    ### RADIOTERAPIA ###
    ###################################

    print("-- -- --RADIOTERAPIA-- -- -- ")
    lista_eliminar = ['na',' na']
    [dic['116'].pop(key,None) for key in lista_eliminar]
    dic_16 ={
        'conformal':"922443",
        'conformal 3d':"922443",
    }
    import radioterapia
    dic['116'].update(dic_16)     
    try:
        __112,__113,__114,__115,__116,__117,__118,__120,__121,__122,__124 = radioterapia._112__131(folios,dic)
    except Exception as e:
        print(e)
        print("falló _112__131")
    print("Salio de radio\n")


    ###################################
    ### DOLOR Y CUIDADOS PALIATIVOS ###
    ###################################
    print("-- -- --Dolor  y Cuidado -- -- -- ")
    def _140__148(folios):
        _140 = None
        _146 = "1"
        patron_fecha = "[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]"
        for folio in folios:
            folio = folio.replace("\n"," ").replace("  "," ").replace("   "," ").replace("  "," ")
            start = folio.find("reg.")
            frag = folio[start:start+60]

            if 'dolor y cuidados' in frag:
                fechas = re.findall(patron_fecha,folio.replace("\n"," ").replace("  "," ").replace("   "," ").replace("  "," "))
                fecha = fechas[0]
                _147 = '-'.join(fecha.split("/")[::-1])
                _140 =  "1"
                _141 = "1"
                _148 = "80010054401"
                break
        if _140 == "1":
            for folio in folios:
                folio = folio.replace("\n"," ").replace("  "," ").replace("   "," ").replace("  "," ")
                start = folio.find("reg.")
                frag = folio[start:start+60]
                if 'psicologia' in frag:
                    _146 = "1" # 1


        if _140 == None:
            _140 = "2"
            _141 = "2"
            _146 = "2" # 2
            _147 =  "1845-01-01"
            _148 = "98" # "80010054401"
        _142,_143,_144,_145 = ("2","2","2","2")
  
        return _140,_141,_142,_143,_144,_145,_146,_147,_148
    try:
        __140,__141,__142,__143,__144,__145,__146,__147,__148 = _140__148(folios)
    except:
        print("falló _140__148")
    print('__140:', 140)
    print('__141:', 141)
    print('__142:', 142)
    print('__143:', 143)
    print('__144:', 144)
    print('__145:', 145)
    print('__146:', 146)
    print('__147:', 147)
    print('__148:', 148)
    print(__140,__141,__142,__143,__144,__145,__146,__147,__148)
    print("\n")

   
    ###################
    ### NUTRICION ###
    ###################
    print("-- -- -- Nutricion -- -- -- ")
    def _152__156(folios):
        _152 = "98"
        patron_fecha = "[0-9][0-9]/[0-9][0-9]/[0-9][0-9][0-9][0-9]"
        for folio in folios:
            folio = folio.replace("\n"," ").replace("  "," ").replace("   "," ").replace("  "," ")
            start = folio.find("reg.")
            frag = folio[start:start+60]
            if 'nutricion' in frag:
                start = folio.find('fecha') + 6
                end = start + 10
                _152 = "1"
                fechas = re.findall(patron_fecha,folio.replace("\n"," ").replace("  "," ").replace("   "," ").replace("  "," "))
                fecha = fechas[0]
                _153 = '-'.join(fecha.split("/")[::-1])
                _155 = "1"
                _154 = "80010054401"
                break
        if _152 == "1":
            here = []
            for folio in folios:
                folio = folio.replace("\n"," ").replace("  "," ").replace("   "," ").replace("  "," ")
                start = folio.find("reg.")
                frag = folio[start:start+60]
                if 'nutricion' in frag:
                    here.append(folio)
            enteral = False 
            parenteral = False
            for aqui in here[::-1]:
                if 'parenteral' in aqui:
                    parenteral = True
                    _155 = '2'
                    break
                elif ' enteral' in aqui:
                    enteral = True
                    _155 = '1'
                    break
                else :
                    _155 = '4'

            for aqui in here[::-1]:
                if 'parenteral' in aqui:
                    parenteral = True
                if ' enteral' in aqui:
                    enteral = True
                if parenteral and enteral:
                    _155 = '3'
                    break
                else:
                    parenteral = False
                    enteral = False
            print(_155)           
        if _152 == "98":
            _152 = "98"
            _155 = "4"
            _154 = "98"
            _153 = "1845-01-01"
        _156 = "98"
        return _152,_153,_154,_155,_156
        
    # try:
    __152,__153,__154,__155,__156 = _152__156(folios)
    print(__152,__153,__154,__155,__156)
    # except Exception as e:
    #     print("falló _152__156")
    #     print(e)

    ###################################### 
    ### RESULTADO FINAL DE LA ATENCIÓN ###
    ######################################
    import resultado_final


    # try:
    __157,__158,__159,__160,__161,__162,__163,__164,__165,__166,__41 = resultado_final._157__166(folios,dic,__45,__100,__112,__6,__5,__76,__98,__121,__130)
    print(f"""
    __157 = {__157}
    __158 = {__158}
    __159 = {__159}
    __160 = {__160}
    __161 = {__161}
    __162 = {__162}
    __163 = {__163}
    __164 = {__164}
    __165 = {__165}
    __166 = {__166}
    """)
    # except Exception as e:
    #     print(e)
    #     print("falló _157__166")
        

    __11 = Eps
    __166 = Fcorte
    __16 = "1800-01-01"



    variables_final = (__1 ,__2 ,__3 ,__4 ,__5 ,__6 ,__7 ,__8 ,__9 ,__10 ,__11 ,__12 ,__13 ,__14 ,__15 ,__16 ,__17 ,__18 ,__19 ,__20 ,__21 ,__22 ,__23 ,__24 ,__25 ,__26 ,__27 ,__28 ,__29 ,__30 ,__31 ,__32 ,__33 ,__34 ,__35 ,__36 ,__37 ,__38 ,__39 ,__40 ,__41 ,__42 ,__43 ,__44 ,__45 ,__46 ,__47 ,__48 ,__49 ,__50 ,__51 ,__52 ,__53 ,__54 ,__55 ,__56 ,__57 ,__58 ,__59 ,__60 ,__61 ,__62 ,__63 ,__64 ,__65 ,__66 ,__67 ,__68 ,__69 ,__70 ,__71 ,__72 ,__73 ,__74 ,__75 ,__76 ,__77 ,__78 ,__79 ,__80 ,__81 ,__82 ,__83 ,__84 ,__85 ,__86 ,__87 ,__88 ,__89 ,__90 ,__91 ,__92 ,__93 ,__94 ,__95 ,__96 ,__97 ,__98 ,__99 ,__100 ,__101 ,__102 ,__103 ,__104 ,__105 ,__106 ,__107 ,__108 ,__109 ,__110 ,__111 ,__112 ,__113 ,__114 ,__115 ,__116 ,__117 ,__118 ,__119 ,__120 ,__121 ,__122 ,__123 ,__124 ,__125 ,__126 ,__127 ,__128 ,__129 ,__130 ,__131 ,__132 ,__133 ,__134 ,__135 ,__136 ,__137 ,__138 ,__139 ,__140 ,__141 ,__142 ,__143 ,__144 ,__145 ,__146 ,__147 ,__148 ,__149 ,__150 ,__151 ,__152 ,__153 ,__154 ,__155 ,__156 ,__157 ,__158 ,__159 ,__160 ,__161 ,__162 ,__163 ,__164 ,__165 ,__166) 
    print('variables_final:', variables_final)
    #########################################
    ### TRABAJANDO CON WORKBOOK DE EXCELL ###
    #########################################

    wb = load_workbook(filename="prueba2.xlsx")
    #wb.create_sheet('CAC',0)
    ws = wb['CAC']

    # print('wb:', wb)
    # print('ws:', ws)

    for i in range(16,167):
        try:
            ws.cell(row=row,column=i,value=variables_final[i-1])
        except:
            ws.cell(row=row,column=i,value=" ")

    wb.save("prueba2.xlsx")

    
    # print('wb (2da parte):', wb)
    # print('ws (2da parte):', ws)



    #########################################
    ### COMIENZO DE  FUNCION COMO MAIN.PY ###
    #########################################

if __name__ == '__main__':
    hcs = glob.glob('*.pdf')
    for hc in hcs:
        pdf_to_txt(hc[:-4])

        continue
    
    pacientes = glob.glob('*32672854*.txt')
    pacientes = glob.glob('HISTORIA*.txt')
    print(f"La cantidad de pacientes es {len(pacientes)}")

    print(' ')
    row = 7
    
    f = '2022-09-30'
    eps = 'Testing'
    print('pacientes:', pacientes)
    for paciente in pacientes:
        # try:
        print('paciente_procesador2:', paciente)
        lon = len(f"== == == == == {paciente} / {row} == == == == ==")
        print("="*lon)
        print(f'== == == == == {paciente} / {row} == == == == ==')
        print("="*lon)
        main(paciente, row, f, eps)
        # except Exception as e:
        #     print("el error es:", e)

        row = row + 1
        for i in range(3):
            print(' ')
            
    print('-- -- -- -- PROCESO TERMINADO -- -- -- -- ')
    print("desde bnndn5") 
    